#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
智慧停车管理平台 - 采购询价功能清单深度分析脚本
"""

import sys
import json
from pathlib import Path
from openpyxl import load_workbook

def deep_analyze_excel(file_path):
    """深度分析 Excel 文件的所有内容"""
    print(f"\n{'='*80}")
    print(f"深度分析 Excel 文件：{file_path}")
    print('='*80)
    
    try:
        # 加载工作簿 - 使用不同的参数组合尝试
        wb = load_workbook(filename=file_path, read_only=False, data_only=True)
        sheet_names = wb.sheetnames
        
        print(f"\n📊 工作表数量：{len(sheet_names)}")
        print(f"📋 工作表名称：{sheet_names}")
        
        all_data = {}
        
        for idx, sheet_name in enumerate(sheet_names):
            print(f"\n{'='*80}")
            print(f"工作表 {idx+1}: 【{sheet_name}】")
            print('='*80)
            
            ws = wb[sheet_name]
            
            # 获取最大行和列
            max_row = ws.max_row
            max_col = ws.max_column
            
            print(f"  最大行数：{max_row}")
            print(f"  最大列数：{max_col}")
            
            # 获取所有有值的单元格
            all_rows = []
            for row_idx in range(1, max_row + 1):
                row_data = []
                for col_idx in range(1, max_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    value = cell.value
                    if value is not None:
                        row_data.append(str(value).strip())
                    else:
                        row_data.append('')
                
                # 只添加非空行
                if any(cell.strip() for cell in row_data if isinstance(cell, str)):
                    all_rows.append(row_data)
            
            print(f"\n  非空行数：{len(all_rows)}")
            
            if all_rows:
                print(f"\n  📋 完整数据内容:")
                print(f"  {'-'*78}")
                
                # 显示所有行
                for i, row in enumerate(all_rows, 1):
                    non_empty_cells = [cell for cell in row if cell.strip()]
                    if non_empty_cells:
                        row_str = ' | '.join(row)
                        print(f"  行{i}: {row_str}")
                
                print(f"  {'-'*78}")
                
                # 智能识别数据结构
                if len(all_rows) >= 2:
                    print(f"\n  🔍 数据结构分析:")
                    
                    # 假设第一行是标题
                    headers = all_rows[0] if all_rows[0] else []
                    print(f"  可能的标题行：{headers}")
                    
                    # 统计每列的数据
                    if len(headers) > 1:
                        print(f"\n  各列数据统计:")
                        for col_idx, header in enumerate(headers, 1):
                            if col_idx <= len(all_rows[0]):
                                col_values = [row[col_idx-1] for row in all_rows[1:] 
                                             if col_idx-1 < len(row) and row[col_idx-1].strip()]
                                print(f"    列{col_idx} [{header}]: {len(col_values)} 条数据")
                                if col_values and len(col_values) <= 5:
                                    print(f"      值：{col_values}")
            
            all_data[sheet_name] = all_rows
        
        return all_data
        
    except Exception as e:
        print(f"\n❌ 分析失败：{str(e)}")
        import traceback
        traceback.print_exc()
        return None

def generate_detailed_report(data):
    """生成详细的需求分析报告"""
    print(f"\n\n{'='*80}")
    print("📋 智慧停车管理平台 - 需求功能分析报告")
    print('='*80)
    
    if not data:
        print("\n⚠️ 无数据可分析")
        return
    
    total_features = 0
    
    for sheet_name, rows in data.items():
        print(f"\n{'─'*80}")
        print(f"【{sheet_name}】")
        print('─'*80)
        
        if not rows:
            print("  (无数据)")
            continue
        
        # 如果有多列，可能是结构化数据
        if len(rows) > 0 and len(rows[0]) > 1:
            print(f"\n  📊 功能清单详情:")
            
            # 提取标题
            headers = rows[0] if rows else []
            
            # 找出关键列（可能包含"功能"、"需求"、"模块"等关键词）
            key_columns = []
            for idx, header in enumerate(headers):
                if any(kw in header.lower() for kw in ['功能', '需求', '模块', '名称', '描述', '优先级']):
                    key_columns.append(idx)
            
            print(f"\n  检测到的关键列：{[headers[i] for i in key_columns]}")
            
            # 提取功能项
            features = []
            for row_idx, row in enumerate(rows[1:], 2):  # 跳过标题行
                feature_info = {}
                for col_idx in key_columns:
                    if col_idx < len(row):
                        feature_info[headers[col_idx]] = row[col_idx]
                
                if feature_info:
                    features.append(feature_info)
                    total_features += 1
            
            if features:
                print(f"\n  共识别出 {len(features)} 个功能点:")
                for i, feat in enumerate(features[:30], 1):  # 显示前 30 个
                    content = ' | '.join(f"{k}: {v}" for k, v in feat.items() if v.strip())
                    if content.strip():
                        print(f"    {i}. {content}")
                
                if len(features) > 30:
                    print(f"    ... 还有 {len(features) - 30} 个功能点")
        
        elif len(rows) > 0:
            # 单列数据，可能是文本列表
            print(f"\n  📝 文本内容:")
            for i, row in enumerate(rows, 1):
                content = row[0] if row else ''
                if content.strip():
                    print(f"    {i}. {content}")
    
    print(f"\n{'='*80}")
    print(f"📈 总结：共识别出 {total_features} 个功能点")
    print('='*80)

def main():
    excel_file = r"f:\code\AICoding\coding\AICoding\0203智慧停车管理平台_采购询价功能清单.xlsx"
    
    print("="*80)
    print("智慧停车管理平台 - 采购询价功能清单深度分析工具")
    print("="*80)
    print(f"\nPython 版本：{sys.version}")
    print(f"文件路径：{excel_file}")
    
    # 检查文件是否存在
    if not Path(excel_file).exists():
        print(f"\n❌ 文件不存在：{excel_file}")
        return
    
    print(f"\n✅ 文件存在，开始深度分析...")
    
    # 执行深度分析
    data = deep_analyze_excel(excel_file)
    
    # 生成报告
    if data:
        generate_detailed_report(data)
        
        # 保存详细结果到 JSON
        output_file = r"f:\code\AICoding\coding\AICoding\excel_detailed_analysis.json"
        try:
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            print(f"\n💾 详细分析结果已保存到：{output_file}")
        except Exception as e:
            print(f"\n⚠️ 保存 JSON 失败：{e}")
    
    print(f"\n{'='*80}")
    print("分析完成！")
    print('='*80)

if __name__ == "__main__":
    main()
