#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
智慧停车管理平台 - 采购询价功能清单分析脚本
"""

import sys
import json
from pathlib import Path

# 尝试导入 openpyxl
try:
    import openpyxl
    from openpyxl import load_workbook
    print("✓ openpyxl 库可用")
except ImportError:
    print("✗ openpyxl 库未安装，尝试使用 pandas...")
    try:
        import pandas as pd
        print("✓ pandas 库可用")
    except ImportError:
        print("✗ pandas 库也未安装")
        print("\n请安装必要的库:")
        print("C:\\Python314\\python.exe -m pip install openpyxl pandas")
        sys.exit(1)

def analyze_with_openpyxl(file_path):
    """使用 openpyxl 分析 Excel 文件"""
    print(f"\n{'='*60}")
    print(f"使用 openpyxl 读取文件：{file_path}")
    print('='*60)
    
    try:
        # 加载工作簿
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        print(f"\n📊 工作表数量：{len(sheet_names)}")
        print(f"📋 工作表名称：{sheet_names}")
        
        all_data = {}
        
        for idx, sheet_name in enumerate(sheet_names):
            print(f"\n{'─'*60}")
            print(f"工作表 {idx+1}: {sheet_name}")
            print('─'*60)
            
            ws = wb[sheet_name]
            
            # 获取所有行
            rows = list(ws.iter_rows(values_only=True))
            
            if not rows:
                print("  (空工作表)")
                continue
            
            print(f"  总行数：{len(rows)}")
            
            # 分析前 5 行作为示例
            print(f"\n  前 5 行数据示例:")
            for i, row in enumerate(rows[:5]):
                if any(cell is not None for cell in row):
                    print(f"    行{i+1}: {row}")
            
            # 检测列名
            if rows and rows[0]:
                headers = [str(cell) if cell is not None else f"列{j+1}" 
                          for j, cell in enumerate(rows[0])]
                print(f"\n  检测到列名 ({len(headers)}列):")
                for j, header in enumerate(headers):
                    print(f"    {j+1}. {header}")
            
            # 存储数据用于后续分析
            all_data[sheet_name] = rows
        
        return all_data
        
    except Exception as e:
        print(f"\n❌ 读取失败：{str(e)}")
        import traceback
        traceback.print_exc()
        return None

def analyze_with_pandas(file_path):
    """使用 pandas 分析 Excel 文件"""
    print(f"\n{'='*60}")
    print(f"使用 pandas 读取文件：{file_path}")
    print('='*60)
    
    try:
        # 读取所有工作表
        all_sheets = pd.read_excel(file_path, sheet_name=None, header=None)
        
        print(f"\n📊 工作表数量：{len(all_sheets)}")
        
        for sheet_name, df in all_sheets.items():
            print(f"\n{'─'*60}")
            print(f"工作表：{sheet_name}")
            print('─'*60)
            print(f"  形状：{df.shape[0]} 行 × {df.shape[1]} 列")
            print(f"\n  前 5 行数据:")
            print(df.head().to_string())
            
        return all_sheets
        
    except Exception as e:
        print(f"\n❌ 读取失败：{str(e)}")
        return None

def generate_requirements_analysis(data):
    """生成需求分析报告"""
    print(f"\n\n{'='*60}")
    print("📋 需求功能分析报告")
    print('='*60)
    
    if not data:
        print("无数据可分析")
        return
    
    for sheet_name, rows in data.items():
        if isinstance(rows, list) and len(rows) > 0:
            print(f"\n【{sheet_name}】功能点统计:")
            
            # 简单统计非空行
            non_empty_rows = [r for r in rows if any(cell is not None for cell in r)]
            print(f"  总记录数：{len(non_empty_rows)}")
            
            # 如果第一行是标题，从第二行开始分析
            if len(non_empty_rows) > 1:
                print(f"\n  功能项列表:")
                for i, row in enumerate(non_empty_rows[1:20], 1):  # 只显示前 20 条
                    cells = [str(cell).strip() if cell is not None else '' for cell in row]
                    content = ' | '.join(cells)
                    if content.strip():
                        print(f"    {i}. {content}")

def main():
    # Excel 文件路径
    excel_file = r"f:\code\AICoding\coding\AICoding\0203智慧停车管理平台_采购询价功能清单.xlsx"
    
    print("="*60)
    print("智慧停车管理平台 - 采购询价功能清单分析工具")
    print("="*60)
    print(f"\nPython 版本：{sys.version}")
    print(f"文件路径：{excel_file}")
    
    # 检查文件是否存在
    if not Path(excel_file).exists():
        print(f"\n❌ 文件不存在：{excel_file}")
        return
    
    print(f"\n✅ 文件存在，开始分析...")
    
    # 优先使用 openpyxl
    if 'openpyxl' in sys.modules:
        data = analyze_with_openpyxl(excel_file)
    else:
        data = analyze_with_pandas(excel_file)
    
    # 生成需求分析
    if data:
        generate_requirements_analysis(data)
        
        # 保存分析结果到 JSON
        output_file = r"f:\code\AICoding\coding\AICoding\excel_analysis_result.json"
        try:
            # 转换数据为可序列化格式
            serializable_data = {}
            for key, value in data.items():
                if hasattr(value, 'to_dict'):  # pandas DataFrame
                    serializable_data[key] = value.to_dict()
                else:  # list of rows
                    serializable_data[key] = [
                        [str(cell) if cell is not None else None for cell in row]
                        for row in value
                    ]
            
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(serializable_data, f, ensure_ascii=False, indent=2)
            print(f"\n💾 分析结果已保存到：{output_file}")
        except Exception as e:
            print(f"\n⚠️ 保存 JSON 失败：{e}")
    
    print(f"\n{'='*60}")
    print("分析完成!")
    print('='*60)

if __name__ == "__main__":
    main()
